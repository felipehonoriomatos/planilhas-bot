

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

ARQUIVO_ENTRADA  = "exemplos/vendas_exemplo.xlsx"
ARQUIVO_SAIDA    = "saida/relatorio_formatado.xlsx"
RESUMO_SAIDA     = "saida/resumo.txt"

COLUNA_CATEGORIA = "Produto"  
COLUNA_VALOR     = "Valor"     
COLUNA_DATA      = "Data"      

COR_CABECALHO   = "1F4E79"   
COR_ACIMA_MEDIA = "C6EFCE"   
COR_ABAIXO_MEDIA= "FFC7CE"   
COR_TOTAL       = "FFE699"   


def criar_planilha_exemplo():
    os.makedirs("exemplos", exist_ok=True)
    dados = {
        "Data": [
            "2024-01-05","2024-01-12","2024-01-20",
            "2024-02-03","2024-02-14","2024-02-22",
            "2024-03-01","2024-03-15","2024-03-28",
            "2024-04-10","2024-04-18","2024-04-25",
        ],
        "Produto": [
            "Notebook","Smartphone","Tablet",
            "Notebook","Smartphone","Tablet",
            "Notebook","Smartphone","Tablet",
            "Notebook","Smartphone","Tablet",
        ],
        "Vendedor": [
            "Ana","Bruno","Ana",
            "Carlos","Ana","Bruno",
            "Bruno","Carlos","Ana",
            "Ana","Bruno","Carlos",
        ],
        "Quantidade": [2, 5, 3, 1, 8, 4, 3, 6, 2, 4, 3, 5],
        "Valor":      [
            4500, 3200, 1800,
            4500, 3200, 1800,
            4500, 3200, 1800,
            4500, 3200, 1800,
        ],
    }
    df = pd.DataFrame(dados)
    df["Total"] = df["Quantidade"] * df["Valor"]
    df.to_excel(ARQUIVO_ENTRADA, index=False)
    print(f"Planilha de exemplo criada em: {ARQUIVO_ENTRADA}")


def formatar_cabecalho(ws, num_colunas: int):
    fill  = PatternFill("solid", fgColor=COR_CABECALHO)
    fonte = Font(color="FFFFFF", bold=True, size=11)
    alin  = Alignment(horizontal="center", vertical="center")
    borda = Border(
        bottom=Side(style="medium", color="FFFFFF")
    )
    for col in range(1, num_colunas + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill  = fill
        cell.font  = fonte
        cell.alignment = alin
        cell.border    = borda
    ws.row_dimensions[1].height = 22


def destacar_por_media(ws, col_valor: int, linha_inicio: int, linha_fim: int, media: float):
    fill_verde    = PatternFill("solid", fgColor=COR_ACIMA_MEDIA)
    fill_vermelho = PatternFill("solid", fgColor=COR_ABAIXO_MEDIA)
    for row in range(linha_inicio, linha_fim + 1):
        cell = ws.cell(row=row, column=col_valor)
        try:
            if float(cell.value) >= media:
                cell.fill = fill_verde
            else:
                cell.fill = fill_vermelho
        except (TypeError, ValueError):
            pass


def adicionar_linha_total(ws, df: pd.DataFrame, col_valor: int):
    prox_linha = ws.max_row + 2
    fill  = PatternFill("solid", fgColor=COR_TOTAL)
    fonte = Font(bold=True, size=11)

    ws.cell(row=prox_linha, column=1, value="TOTAL GERAL").font = fonte
    ws.cell(row=prox_linha, column=1).fill = fill

    total = df[COLUNA_VALOR].sum()
    cell_total = ws.cell(row=prox_linha, column=col_valor, value=total)
    cell_total.font = fonte
    cell_total.fill = fill
    cell_total.number_format = 'R$ #,##0.00'

    return prox_linha


def adicionar_grafico(ws, df: pd.DataFrame):
    resumo = df.groupby(COLUNA_CATEGORIA)[COLUNA_VALOR].sum().reset_index()

    aba_grafico = ws.parent.create_sheet("Dados_Grafico")
    aba_grafico.append([COLUNA_CATEGORIA, "Total Vendas"])
    for _, row in resumo.iterrows():
        aba_grafico.append([row[COLUNA_CATEGORIA], row[COLUNA_VALOR]])

    chart = BarChart()
    chart.type  = "col"
    chart.title = "Total de Vendas por Produto"
    chart.y_axis.title = "Valor (R$)"
    chart.x_axis.title = "Produto"
    chart.style = 10
    chart.width  = 18
    chart.height = 12

    n = len(resumo) + 1
    data  = Reference(aba_grafico, min_col=2, min_row=1, max_row=n)
    cats  = Reference(aba_grafico, min_col=1, min_row=2, max_row=n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "I2")


def ajustar_colunas(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 4


def gerar_resumo_texto(df: pd.DataFrame) -> str:
    total   = df[COLUNA_VALOR].sum()
    media   = df[COLUNA_VALOR].mean()
    maior   = df[COLUNA_VALOR].max()
    menor   = df[COLUNA_VALOR].min()
    por_cat = df.groupby(COLUNA_CATEGORIA)[COLUNA_VALOR].sum()

    linhas = [
        "=" * 45,
        f"  RELATÓRIO GERADO EM {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "=" * 45,
        f"\nTotal de registros : {len(df)}",
        f"Soma total          : R$ {total:,.2f}",
        f"Média por registro  : R$ {media:,.2f}",
        f"Maior valor         : R$ {maior:,.2f}",
        f"Menor valor         : R$ {menor:,.2f}",
        f"\n{'─'*45}",
        f"  TOTAIS POR {COLUNA_CATEGORIA.upper()}",
        f"{'─'*45}",
    ]
    for cat, val in por_cat.items():
        linhas.append(f"  {cat:<20} R$ {val:>12,.2f}")

    linhas.append("\n" + "=" * 45)
    return "\n".join(linhas)


def main():
    os.makedirs("saida", exist_ok=True)

    if not os.path.exists(ARQUIVO_ENTRADA):
        criar_planilha_exemplo()

    print(f"\nLendo: {ARQUIVO_ENTRADA}")
    df = pd.read_excel(ARQUIVO_ENTRADA)
    print(f"   → {len(df)} linhas encontradas")

    wb = load_workbook(ARQUIVO_ENTRADA)
    ws = wb.active

    colunas = [c.value for c in ws[1]]
    col_valor_idx = colunas.index(COLUNA_VALOR) + 1

    formatar_cabecalho(ws, len(colunas))

    media = df[COLUNA_VALOR].mean()
    destacar_por_media(ws, col_valor_idx, 2, len(df) + 1, media)

    for row in range(2, len(df) + 2):
        ws.cell(row=row, column=col_valor_idx).number_format = 'R$ #,##0.00'

    adicionar_linha_total(ws, df, col_valor_idx)

    adicionar_grafico(ws, df)

    ajustar_colunas(ws)

    wb.save(ARQUIVO_SAIDA)
    print(f"Relatório Excel salvo em: {ARQUIVO_SAIDA}")

    resumo = gerar_resumo_texto(df)
    print(resumo)
    with open(RESUMO_SAIDA, "w", encoding="utf-8") as f:
        f.write(resumo)
    print(f"Resumo salvo em: {RESUMO_SAIDA}")


if __name__ == "__main__":
    main()
